import openpyxl
import requests
from bs4 import BeautifulSoup

# FUNCTIONS ------------------------------------------------------
def checkJobTitles(jobs):
    acceptedTitles = ["Full Stack", "Fullstack", "Full-Stack" "Software Engineer", "Software Developer", "Recruiter", "Recruitment", "Python", "Application Developer", "Backend Developer", "Back End Developer", "Web Developer"]
    for title in acceptedTitles:
        if title in jobs:
            return True
    return False

def getAllLinksOnPage(response, pgURL):
    soup = BeautifulSoup(response.text, 'html.parser')

    # Find all the links on the page
    jobs = {}
    url = pgURL
    links = soup.findAll("a")
    for link in links:
        # If link contains any of accepted job titles, add it to jobs list
        if checkJobTitles(link.text):
            jobURL = link.get('href')

            if (jobURL.find("https") >= 0):
                url = ""
            else:
                # Removing everything from main url after .com/.io/.ca/etc
                index = url.find("//")
                index = url.find("/", index+2)
                url = url[0:index]

            jobs[link.text] = f"{url}{jobURL}"      # TODO: Validate jobURL -- if not valid, just use original pgURL

    return jobs

# ----------------------------------------------------------------

# ==========================================================================================================================================
#                                 GETTING LIST OF RELEVANT CAREERS FROM EACH COMPANY CAREER PAGE
# ==========================================================================================================================================
# IMPORTANT VARIABLES:
companies = []
jobsList = {}

# STEP 1: Getting a list of companies that I know already have unlisted jobs. Reading these from a local Excel File (companies_url_list.xlsx)

# Open Workbook and then Sheet1 (companiesList)
wb = openpyxl.load_workbook("companies_url_list.xlsx")
companiesList = wb["Sheet1"]

# Traverse through each row in companiesList and get the specific url -- insert into companies
for row in range(1, companiesList.max_row + 1):
    url = companiesList.cell(row, 1)
    companies.append(url.value)


# STEP 2: For each URL in companies (career pages), scrape HTML document
#         Pull out all the links (a tags) and check if they match appropriate job titles
#         If they do, add it to jobs list
# Get the HTML using Requests
for pageURL in companies:
    response = requests.get(pageURL)

    if response.status_code == 200:   # 200 is OK
        jobsList.update(getAllLinksOnPage(response, pageURL))


#for job in jobsList:
 #   print(job)
print(jobsList)
